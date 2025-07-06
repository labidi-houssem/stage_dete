import scan_one_doc
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QFileDialog, QTextBrowser, QLabel
from PyQt5.uic import loadUi
import os
import pytesseract
import pandas as pd
import openpyxl
import re
from PyQt5.QtGui import QPixmap
import requests


def scan_document_with_flask(document_paths):
    # URL of the Flask API for document scanning
    api_url = 'http://127.0.0.1:5000/api/scan_document'

    for document_path in document_paths:
        # Open the document as a file
        with open(document_path, 'rb') as document_file:
            # Make a POST request to the Flask API sending the document
            response = requests.post(
                api_url, files={'document': document_file})

            if response.status_code == 200:
                message = response.json().get('message', 'Document processed successfully.')
                print(message)
            else:
                print("Error during request to Flask API:",
                      response.json().get('message', 'Unknown error'))


def extraire_texte(image):
    texte_extrait = pytesseract.image_to_string(image, lang='eng')
    return texte_extrait


def sauvgardetext(texte):
    # Nom du fichier Excel
    nom_fichier_excel = "text.xlsx"
    # Si le fichier Excel existe déjà, ouvrez-le
    if os.path.exists(nom_fichier_excel):
        classeur = openpyxl.load_workbook(nom_fichier_excel)
        feuille = classeur.active
    else:
        # Sinon, créez un nouveau fichier Excel
        classeur = openpyxl.Workbook()
        feuille = classeur.active
    lignes = texte.split('\n')
    # Trouver la dernière ligne utilisée dans la feuille
    derniere_ligne = feuille.max_row + 1
    # Écrire chaque ligne dans une nouvelle ligne de la feuille
    for i, ligne in enumerate(lignes, start=derniere_ligne):
        feuille.cell(row=i, column=1, value=ligne)
    classeur.save(nom_fichier_excel)
    import subprocess
    subprocess.Popen([nom_fichier_excel], shell=True)


def sauvgarder1(images):
    import cv2
    nom_dossier = "dataset"
    if not os.path.exists(nom_dossier):
        os.makedirs(nom_dossier)
    x = 0
    for image_path in images:
        x += 1
        image = cv2.imread(image_path)
        image_name = str(x)+".jpg"  # Obtenez le nom du fichier de l'image
        # Chemin complet pour la nouvelle image
        new_image_path = os.path.join(nom_dossier, image_name)
        cv2.imwrite(new_image_path, image)
        try:
            extracted_text = extraire_texte(image_path)
        except Exception as e:
            print(None, "Error", str(e))
        # Split the text into lines
        if extracted_text.find("operation") < 0 and extracted_text.find("date") < 0 and extracted_text.find("Debit") < 0 and extracted_text.find("credit") < 0:
            sauvgardetext(extracted_text)
        else:
            lines = extracted_text.split('\n')
            # Initialize lists to store extracted data
            data = []
            for line in lines:
                # Check if the line starts with a date-like pattern
                if re.match(r"\d{2}/\d{2}", line):
                    parts = line.split()
                    if len(parts) >= 3:
                        date = parts[0]
                        operation = " ".join(parts[1:-2])
                        amount = parts[-1].replace(",", ".")
                        if "Débit" in operation:
                            debit = amount
                            credit = "0.00"
                        else:
                            debit = "0.00"
                            credit = amount
                        data.append([date, operation, debit, credit])
            # Create a DataFrame from the extracted data
            columns = ["Date", "Opérations", "Débit", "Credit"]
            new_df = pd.DataFrame(data, columns=columns)
            # Read the existing Excel file into a DataFrame (if it exists)
            try:
                existing_df = pd.read_excel("datasetone.xlsx")
            except FileNotFoundError:
                existing_df = pd.DataFrame(columns=columns)
            # Concatenate the new data with the existing DataFrame
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            # Save the combined DataFrame to the Excel file
            combined_df.to_excel("datasetone.xlsx", index=False)
        import subprocess
        subprocess.Popen(["datasetone.xlsx"], shell=True)


def sauvegarder_images_dialog():
    options = QFileDialog.Options()
    options |= QFileDialog.ReadOnly
    image_paths, _ = QFileDialog.getOpenFileNames(
        None, "Ouvrir des images", "", "Images (*.png *.jpg *.jpeg *.bmp *.gif);;Tous les fichiers (*)", options=options)
    if image_paths:
        sauvgarder1(image_paths)


app = QApplication(sys.argv)

main_window = loadUi('qtversion4.ui')
scan_window = loadUi('scan.ui')
text_browser = scan_window.findChild(QTextBrowser, 'aff')


def open_scan_page():
    scan_window.show()
    main_window.hide()


def return_to_main_page():
    scan_window.hide()
    main_window.show()


def execute_scan():
    scan_window.txt.setText("wait some seconde...")
    QApplication.instance().processEvents()
    callback = scan_one_doc.Callback("210 x 297")
    callback.fn()
    scan_window.txt.setText("Done")
    QApplication.instance().processEvents()
    import subprocess
    subprocess.Popen(["dataset.xlsx"], shell=True)


def execute_scan_multiple():
    scan_window.txt.setText("wait some seconde...")
    QApplication.instance().processEvents()
    callback = scan_one_doc.Callback("210 x 297")
    callback.scan_multiple_docs(int(scan_window.nbr.text()))
    scan_window.txt.setText("Done")
    QApplication.instance().processEvents()


button_open_scan = main_window.findChild(QPushButton, 'btn2')
button_open_scan.clicked.connect(open_scan_page)
button_return_to_main = scan_window.findChild(QPushButton, 'btn3')
one = scan_window.findChild(QPushButton, 'one')
two = scan_window.findChild(QPushButton, 'two')
one.clicked.connect(scan_one_doc.Callback("210 x 297").fn)

one = scan_window.findChild(QPushButton, 'one')
one.clicked.connect(execute_scan)
two.clicked.connect(execute_scan_multiple)
button_return_to_main.clicked.connect(return_to_main_page)
importimg = main_window.findChild(QPushButton, 'btn')
importimg.clicked.connect(sauvegarder_images_dialog)
main_window.show()
sys.exit(app.exec_())
